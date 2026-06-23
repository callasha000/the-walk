from __future__ import annotations

from datetime import date, datetime
import json
from pathlib import Path
import re
import sys
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as error:
    raise SystemExit(
        "openpyxl is required. Run this script with the Codex bundled Python "
        "runtime or install it with `python -m pip install openpyxl`."
    ) from error


ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = ROOT / "Master Modules Matrix Project No. 25106 Rev 13 April 6 (2).xlsx"
OUTPUT = ROOT / "data" / "module-matrix.json"
SHEET_NAME = "Rev 13, April 6"
EXPECTED_MODULE_COUNT = 499
VALID_FABRICATORS = {"EMP", "West Modular"}


def normalize_dimension(value: Any) -> str | None:
    if value in (None, ""):
        return None

    text = str(value).strip()
    replacements = {
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "''": '"',
    }

    for source, replacement in replacements.items():
        text = text.replace(source, replacement)

    text = re.sub(r"(\d)'(?=\d)", r"\1'-", text)
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_date(value: Any) -> str | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, str):
        text = value.strip()
        if text.upper() == "TBD":
            return "TBD"

        for pattern in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, pattern).date().isoformat()
            except ValueError:
                pass

        return text

    return str(value)


def normalize_text(value: Any) -> str | None:
    if value in (None, ""):
        return None

    return str(value).strip()


def normalize_number(value: Any) -> int | float | None:
    if value in (None, ""):
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value) if value.is_integer() else value

    try:
        number = float(str(value).strip())
    except ValueError:
        return None

    return int(number) if number.is_integer() else number


def normalize_fabricator(value: Any) -> str | None:
    text = normalize_text(value)
    if text in VALID_FABRICATORS:
        return text

    return None


def normalize_status(value: Any) -> str | int | float | None:
    if value in (None, ""):
        return None

    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)

    return str(value).strip()


def schedule_status(row: tuple[Any, ...], status: int, start: int, due: int) -> dict[str, Any]:
    return {
        "status": normalize_status(row[status]),
        "startDate": normalize_date(row[start]),
        "dueDate": normalize_date(row[due]),
    }


def main() -> int:
    if not SOURCE_WORKBOOK.exists():
        print(f"Missing source workbook: {SOURCE_WORKBOOK}", file=sys.stderr)
        return 1

    workbook = load_workbook(SOURCE_WORKBOOK, data_only=True, read_only=True)
    worksheet = workbook[SHEET_NAME]
    records: dict[str, dict[str, Any]] = {}

    for row in worksheet.iter_rows(min_row=16, values_only=True):
        module_id = normalize_text(row[1])
        if not module_id or not re.fullmatch(r"M\d+", module_id):
            continue

        line_1_sequence = normalize_number(row[8])
        line_2_sequence = normalize_number(row[9])
        production_line = 1 if line_1_sequence is not None else 2 if line_2_sequence is not None else None
        production_sequence = line_1_sequence if line_1_sequence is not None else line_2_sequence

        records[module_id] = {
            "module": module_id,
            "item": normalize_number(row[0]),
            "tranche": normalize_number(row[2]),
            "modType": normalize_text(row[3]),
            "moduleSerialNumber": normalize_text(row[4]),
            "oversized": normalize_text(row[5]) == "Oversized",
            "dimension": normalize_dimension(row[6]),
            "estimatedWeightLb": normalize_number(row[7]),
            "productionLine": production_line,
            "productionSequence": production_sequence,
            "chassisShopDrawings": {
                **schedule_status(row, 10, 11, 12),
                "requiredApprovalDate": normalize_date(row[13]),
                "revisionStatus": normalize_status(row[14]),
            },
            "moduleShopDrawings": {
                **schedule_status(row, 15, 16, 17),
                "requiredApprovalDate": normalize_date(row[18]),
                "revisionStatus": normalize_status(row[19]),
            },
            "assignedFabricator": normalize_fabricator(row[20]),
            "chassisFabrication": schedule_status(row, 21, 22, 23),
            "moduleFabrication": schedule_status(row, 24, 25, 26),
            "preYardInspection": {
                **schedule_status(row, 27, 28, 29),
                "notes": normalize_text(row[30]),
            },
            "shipping": {
                "status": normalize_status(row[31]),
                "shippingDate": normalize_date(row[32]),
                "arrivalDate": normalize_date(row[33]),
            },
            "yard": {
                "inspectionDate": normalize_date(row[34]),
                "notes": normalize_text(row[35]),
            },
        }

    if len(records) != EXPECTED_MODULE_COUNT:
        print(
            f"Expected {EXPECTED_MODULE_COUNT} module rows, found {len(records)}.",
            file=sys.stderr,
        )
        return 1

    OUTPUT.write_text(
        json.dumps(dict(sorted(records.items(), key=lambda item: int(item[0][1:]))), indent=2)
        + "\n",
        encoding="utf-8",
    )
    print(OUTPUT.relative_to(ROOT))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
